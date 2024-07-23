import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import logging
import os
from datetime import date, datetime, timedelta

import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
from shared.db_utils import close_db_conn, f_get_table

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
)


no_of_days = int(sys.argv[1]) if len(sys.argv) > 1 else 28


def f_folder_creator(date_start, append_timestamp=True):
    date_format = date_start.strftime("%Y%m%b")
    if append_timestamp:
        date_format += datetime.now().strftime("_%H%M%S")
    if not os.path.exists(date_format):
        os.makedirs(date_format)
    return date_format


def main():
    # Create a message for date range selection
    msg0 = (
        "Selecting date range for last month: ("
        + (datetime.now() - relativedelta(months=1)).strftime("%B %Y")
        + ")"
    )
    logging.info(msg0)

    # Calculate the last day of the previous month
    last_day_prev_mth = date.today().replace(day=1) - timedelta(days=1)

    # Calculate the start and end dates for the selected month
    date_start = last_day_prev_mth.replace(day=1)
    date_end = last_day_prev_mth

    target_folder = f_folder_creator(date_start)

    # grab all STK meters using combined meter table with site, proj, cust, STK is cust id 31
    query = """
        select address, meter_name, excel_name, meter.meter_id, meter_type 
        from meter 
        left join site on site.site_id = meter.site_id
        left join proj on proj.proj_id = site.proj_id
        left join customer on customer.cust_id = proj.cust_id
        left join meter_billing on meter_billing.meter_id = meter.meter_id
        where customer.cust_id = 31 and billing=1
    """
    STK_comb_meter_table = f_get_table(query)
    STK_meter_ids = STK_comb_meter_table["meter_id"]
    STK_meter_ids_str = ", ".join(map(str, STK_meter_ids))

    # grab comb energy table
    query = f"select meter_id,date,energy from energy_patched where meter_id in ({STK_meter_ids_str}) and date >= '{date_start}' and date <= '{date_end}'"
    comb_energy_table = f_get_table(query)

    # create df to store number of days of data each meter has
    data_availability_df = pd.DataFrame(
        index=range(len(STK_comb_meter_table)), columns=["meter_name", "data_points"]
    )

    for i in range(len(STK_meter_ids)):
        # populate data_availability_df
        data_points = comb_energy_table[
            comb_energy_table["meter_id"] == STK_meter_ids[i]
        ].shape[0]
        meter_name = STK_comb_meter_table[
            STK_comb_meter_table["meter_id"] == STK_meter_ids[i]
        ]["meter_name"].values[0]
        data_availability_df.loc[i, "meter_name"] = meter_name
        data_availability_df.loc[i, "data_points"] = data_points
        # write log if any lostcomms found
        if data_points < no_of_days:
            msg1 = f"{meter_name} only has {data_points} data points"
            with open("lostcomms_detected_log.log", "a") as file:
                file.write(msg1 + "\n")

    # write data_availability_df as csv
    data_availability_df.to_csv("data_availability.csv")

    comb_df = pd.merge(
        STK_comb_meter_table, comb_energy_table, on="meter_id", how="left"
    )
    comb_df = comb_df[comb_df["energy"].notna()]
    comb_df["date"] = pd.to_datetime(comb_df["date"], format="%Y-%m-%d")
    comb_df = comb_df.sort_values(by="date")

    summary_df = comb_df.pivot_table(
        index="address", columns="meter_type", values="energy", aggfunc="sum"
    ).reset_index()
    summary_df = summary_df.fillna(0)
    summary_df["consumption"] = (
        summary_df["Solar Generation"] - summary_df["Solar Export"]
    )
    summary_df.columns = ["Site", "Generation(kWh)", "Export(kWh)", "Consumption(kWh)"]

    chinbee_final_df = (
        comb_df[comb_df["address"] == "15 Chin Bee Drive"][
            ["excel_name", "date", "energy"]
        ]
        .pivot(index="date", columns="excel_name", values="energy")
        .reset_index()
    )
    tuas16_final_df = (
        comb_df[comb_df["address"] == "16 Tuas Ave 7"][["excel_name", "date", "energy"]]
        .pivot(index="date", columns="excel_name", values="energy")
        .reset_index()
    )
    jbl249_final_df = (
        comb_df[comb_df["address"] == "249 Jalan Boon Lay"][
            ["excel_name", "date", "energy"]
        ]
        .pivot(index="date", columns="excel_name", values="energy")
        .reset_index()
    )
    benoi16_final_df = (
        comb_df[comb_df["address"] == "16 Benoi Crescent"][
            ["excel_name", "date", "energy"]
        ]
        .pivot(index="date", columns="excel_name", values="energy")
        .reset_index()
    )

    filename = f"STK-{date_start.strftime('%b%Y')}-MonthlyReport-GwExport.xlsx"
    full_path = os.path.join(target_folder, filename)

    with pd.ExcelWriter(full_path) as writer:
        summary_df.to_excel(writer, sheet_name="Summary_Table", index=False)
        chinbee_final_df.to_excel(writer, sheet_name="15_ChinBeeDrive", index=False)
        tuas16_final_df.to_excel(writer, sheet_name="16_TuasAve7", index=False)
        jbl249_final_df.to_excel(writer, sheet_name="249_JalanBoonLay", index=False)
        benoi16_final_df.to_excel(writer, sheet_name="16_BenoiCres", index=False)

    wb = openpyxl.load_workbook(full_path)
    sheets = wb.sheetnames
    for i, sheet_name in enumerate(sheets):
        sheet = wb[sheet_name]
        if i == 0:
            sheet_max_col = len(summary_df.columns)
        elif i == 1:
            sheet_max_col = len(chinbee_final_df.columns)
        elif i == 2:
            sheet_max_col = len(tuas16_final_df.columns)
        elif i == 3:
            sheet_max_col = len(jbl249_final_df.columns)
        elif i == 4:
            sheet_max_col = len(benoi16_final_df.columns)

        for col_index in range(1, sheet_max_col + 1):
            sheet.column_dimensions[
                openpyxl.utils.get_column_letter(col_index)
            ].auto_size = True
    wb.save(full_path)


if __name__ == "__main__":  # pragma: no cover
    try:
        main()
    except Exception as e:
        logging.error(e)
    finally:
        close_db_conn()
